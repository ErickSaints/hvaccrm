"""
Extract CATALOGO_PRECIOS_UNITARIOS_HVAC_CDMX_2026.xlsx into structured JSON
for import into the Pricebook.
"""
import json
import re
import os
from openpyxl import load_workbook

XLSX_PATH = os.path.expanduser(r'~\OneDrive\Desktop\CATALOGO_PRECIOS_UNITARIOS_HVAC_CDMX_2026.xlsx')
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), 'catalog_import.json')

def parse_price(val):
    if val is None:
        return None
    s = str(val).strip()
    s = s.replace(',', '').replace('$', '').replace('~', '').replace(' ', '')
    s = s.replace('*', '').replace('?', '')
    try:
        return float(s)
    except ValueError:
        return None

def parse_price_range(val):
    """Try to extract (low, high) or (single, None) from a price string."""
    if val is None:
        return None, None
    s = str(val).strip()
    s = s.replace(',', '').replace('$', '').replace('~', '').replace('*', '').replace('?', '').strip()
    # Range pattern
    m = re.match(r'^([\d,.]+)\s*[-–—]\s*([\d,.]+)$', s)
    if m:
        low = parse_price(m.group(1))
        high = parse_price(m.group(2))
        return low, high
    single = parse_price(s)
    return single, None

def is_section_header(row_values):
    """Check if a row is a section header (all caps, no price)."""
    text = str(row_values[0] or '').strip()
    if not text:
        return False
    # All caps, longer than 3 chars, and rest of row has no numbers
    if text.isupper() and len(text) > 2:
        numeric_count = sum(1 for v in row_values[1:] if parse_price(v) is not None)
        return numeric_count == 0
    return False

def clean_name(name):
    if not name:
        return ''
    s = str(name).strip()
    if s.isupper():
        s = s.title()
    return s

# ── Sheet configurations ───────────────────────────────────────────────────

SHEET_CONFIGS = {
    '1. Mantto Preventivo': {
        'unit': 'servicio',
        'col_map': {'name': 0, 'capacidad': 1, 'price': 2, 'horas': 3, 'frecuencia': 4, 'fuente': 5},
    },
    '2. Mantto Correctivo': {
        'unit': 'servicio',
        'col_map': {'name': 0, 'price': 1, 'tiempo': 2, 'fuente': 3},
    },
    '3. Reemplazo Componentes': {
        'unit': 'pza',
        'col_map': {'name': 0, 'price': 1, 'notas': 2, 'fuente': 3},
    },
    '4. Gas Refrigerante': {
        'unit': 'kg',
        'col_map': {'name': 0, 'presentacion': 1, 'price_gas': 2, 'price_kg': 3, 'price_service_1': 4, 'price_service_2': 5, 'fuente': 6},
    },
    '5. Instalación Equipos': {
        'unit': 'servicio',
        'col_map': {'name': 0, 'capacidad': 1, 'price_mo': 2, 'price_preinst': 3, 'price_equipo': 4, 'price_total': 5, 'tiempo': 6, 'fuente': 7},
    },
    '6. Chiller (APU Reales)': {
        'unit': 'servicio',
        'col_map': {'name': 0, 'chiller_100': 1, 'chiller_130': 2, 'chiller_800': 3, 'sistema_completo': 4},
    },
    '7. Ductería': {
        'unit': None,
        'col_map': {'name': 0, 'unidad': 1, 'price_cliente': 2, 'costo_directo': 3, 'notas': 4, 'fuente': 5},
    },
    '8. Instalación Eléctrica': {
        'unit': 'servicio',
        'col_map': {'name': 0, 'price': 1, 'fuente': 2},
    },
    '9. Igualas y Pólizas': {
        'unit': 'servicio',
        'col_map': {'name': 0, 'price_mensual': 1, 'price_trimestral': 2, 'price_semestral': 3, 'price_anual': 4, 'notas': 5},
    },
    '10. Recargos y Factores': {
        'unit': None,
        'col_map': {'name': 0, 'factor': 1, 'aplica': 2, 'notas': 3},
        'skip': True,  # Reference data, not pricebook items
    },
    '11. Jornales y MO': {
        'unit': 'jornal',
        'col_map': {'name': 0, 'salario': 1, 'con_prestaciones': 2, 'cuadrilla': 3, 'costo_cuadrilla': 4},
    },
    '12. Centralizado Completo': {
        'unit': 'obra',
        'col_map': {'name': 0, 'price_central': 1, 'price_conductos': 2, 'price_5t': 3, 'fuente': 4},
    },
}

# ── Read Excel ──────────────────────────────────────────────────────────────

wb = load_workbook(XLSX_PATH, data_only=True)

categories = []

for sheet_name in wb.sheetnames:
    config = SHEET_CONFIGS.get(sheet_name)
    if config is None:
        print(f"[WARN] Unknown sheet: {sheet_name}, skipping")
        continue
    if config.get('skip'):
        print(f"[SKIP] Reference sheet: {sheet_name}")
        continue

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header
    col_map = config['col_map']

    cat_name = sheet_name.split('. ', 1)[1] if '. ' in sheet_name else sheet_name
    category_items = []
    current_subcategory = None

    for row in rows:
        values = list(row)
        if not values or not values[0]:
            continue
        name_raw = str(values[0]).strip() if values[0] else ''
        if not name_raw:
            continue

        # Check for section header row
        if is_section_header(values):
            current_subcategory = name_raw.title()
            continue

        # Build item
        name = clean_name(name_raw)
        description_parts = []
        prices = {}  # Track all prices found

        # Standard fields
        name_field = col_map.get('name', 0)

        for key, col_idx in col_map.items():
            if key == 'name' or col_idx >= len(values):
                continue
            val = values[col_idx]
            if val is None or str(val).strip() == '' or str(val).strip() == '—':
                continue
            val_str = str(val).strip()

            # Parse price fields
            if key.startswith('price') or key in ('salario', 'con_prestaciones', 'costo_cuadrilla',
                                                    'costo_directo', 'chiller_100', 'chiller_130',
                                                    'chiller_800', 'sistema_completo'):
                low, high = parse_price_range(val_str)
                if low is not None:
                    prices[key] = (low, high)
                else:
                    description_parts.append(f"{key.replace('_', ' ').title()}: {val_str}")
            elif key in ('capacidad', 'horas', 'frecuencia', 'tiempo', 'notas', 'fuente', 'unidad', 'aplica', 'factor', 'presentacion'):
                description_parts.append(f"{key.replace('_', ' ').title()}: {val_str}")
            else:
                description_parts.append(f"{key.replace('_', ' ').title()}: {val_str}")

        # Determine unit
        unit = config.get('unit')
        if not unit and 'unidad' in col_map:
            unit_val = values[col_map['unidad']]
            if unit_val:
                unit = str(unit_val).strip().lower()
        if not unit:
            unit = 'pza'

        # Add capacity/size context to name if available for uniqueness
        capacidad = None
        if 'capacidad' in col_map and col_map['capacidad'] < len(values):
            cap_val = values[col_map['capacidad']]
            if cap_val:
                capacidad = str(cap_val).strip()
        if capacidad and len(capacidad) < 30 and not name.lower().startswith(str(capacidad).lower()):
            full_name = f"{name} ({capacidad})"
        else:
            full_name = name

        # Build description
        desc_lines = []
        if current_subcategory:
            desc_lines.append(f"Categoría: {current_subcategory}")
        for part in description_parts:
            # Skip redundant/non-informative parts
            if 'Fuente' in part and not any(p.startswith('Fuente') for p in desc_lines):
                desc_lines.append(part)
            elif not part.startswith('Fuente'):
                desc_lines.append(part)
        desc = ' | '.join(desc_lines)

        item = {
            'name': full_name,
            'description': desc,
            'unit': unit,
            'prices': prices,
            'source_sheet': sheet_name,
            'subcategory': current_subcategory,
        }

        # For ducteria, capture costPrice separately
        if sheet_name == '7. Ductería' and 'costo_directo' in prices:
            item['costPrice'] = prices['costo_directo'][0]

        category_items.append(item)

    categories.append({
        'name': cat_name,
        'items': category_items,
    })

    print(f"[OK] {sheet_name}: {len(category_items)} items")

# ── Generate import JSON ───────────────────────────────────────────────────

output = []
for cat in categories:
    for item in cat['items']:
        prices = item['prices']
        good_price = None
        better_price = None
        best_price = None
        cost_price = item.get('costPrice')

        # Map prices to Good/Better/Best tiers
        sheet = item['source_sheet']

        all_price_vals = []  # collect all (low, high, label) tuples
        for key, (low, high) in prices.items():
            if low is not None:
                all_price_vals.append((low, key))
            if high is not None and high != low:
                all_price_vals.append((high, key))

        if sheet == '5. Instalación Equipos':
            # Map MO=good, Preinst=better, Total=best
            low = prices.get('price_mo')
            mid = prices.get('price_preinst')
            high = prices.get('price_total')
            if low:
                good_price = low[0]
            if mid:
                better_price = mid[0]
            if high:
                best_price = high[1] if high[1] else high[0]
        elif sheet == '4. Gas Refrigerante':
            # Map gas price=good, kg price=better, service=best
            pg = prices.get('price_gas')
            pk = prices.get('price_kg')
            ps = prices.get('price_service_1')
            if pg:
                good_price = pg[0]
            if pk:
                better_price = pk[0]
            if ps:
                best_price = ps[1] if ps[1] else ps[0]
        elif sheet == '7. Ductería':
            low = prices.get('price_cliente')
            cost = prices.get('costo_directo')
            if low:
                good_price = low[0]
                best_price = low[1] if low[1] else None
            if cost:
                cost_price = cost[0]
        elif sheet == '9. Igualas y Pólizas':
            low = prices.get('price_mensual')
            high = prices.get('price_anual')
            if low:
                good_price = low[0]
            if high:
                best_price = high[1] if high[1] else high[0]
        elif sheet == '11. Jornales y MO':
            low = prices.get('salario')
            high = prices.get('costo_cuadrilla')
            if low:
                good_price = low[0]
            if high:
                best_price = high[1] if high[1] else high[0]
        else:
            # Generic: sort all numeric values, assign to tiers
            nums = sorted(set(v for v, _ in all_price_vals if v is not None))
            if len(nums) >= 1:
                good_price = nums[0]
            if len(nums) >= 2:
                best_price = nums[-1]
            if len(nums) >= 3:
                better_price = nums[len(nums)//2]

        entry = {
            'name': item['name'],
            'description': item['description'],
            'unit': item['unit'],
            'category': cat['name'],
            'goodPrice': round(good_price, 2) if good_price else None,
            'betterPrice': round(better_price, 2) if better_price else None,
            'bestPrice': round(best_price, 2) if best_price else None,
            'costPrice': round(cost_price, 2) if cost_price else None,
        }
        output.append(entry)

with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False, indent=2)

print(f"\n[OK] Total items extracted: {len(output)}")
print(f"[FILE] Output: {OUTPUT_PATH}")

# Show summary by category
from collections import Counter
cat_counts = Counter(e['category'] for e in output)
for cat_name, count in sorted(cat_counts.items()):
    with_prices = sum(1 for e in output if e['category'] == cat_name and e['goodPrice'] is not None)
    print(f"   {cat_name}: {count} items ({with_prices} con precio)")
